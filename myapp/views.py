from django.shortcuts import render
import openpyxl
from pycoingecko import CoinGeckoAPI


def index(request):
    if "GET" == request.method:
        return render(request, "myapp/index.html", {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        # print(worksheet["A1"].value)

        excel_data_list = []
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            for cell in row:
                excel_data_list.append(str(cell.value))
                print(cell.value)
        excel_data = sorted(set([x.lower() for x in excel_data_list]))
        print(excel_data)

        cg = CoinGeckoAPI()
        data_prices = cg.get_price(ids=excel_data_list, vs_currencies="usd")
        print(data_prices)

        price_list = [
            data_prices[i]["usd"] if i in data_prices.keys() else "wrong spell"
            for i in excel_data
        ]
        print(price_list)

        final_list = list(zip(excel_data, price_list))

        return render(request, "myapp/index.html", {"final_list": final_list})
